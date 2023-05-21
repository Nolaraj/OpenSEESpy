import openpyxl as op
from itertools import combinations

wb = op.load_workbook("Paper.xlsx")
ws = wb["Classified"]
ws2 = wb["Category2"]
ws3 = wb["Category3"]

referencesNO = 52


max_row = ws.max_row
print(max_row)

citation_list = ["[" + str(x) + "]" for x in range(1, referencesNO + 1)]
titles = []
citation_titles = {}

for index, citation_NO in enumerate(citation_list):
    i = index + 1

    parameters = []

    for row in range(6, max_row + 1):
        cell_Value = ws.cell(column = 2, row = row).value

        if  citation_NO == cell_Value:

                j = row
                title = ws.cell(column = 2, row = j).value
                while (title != None and title in citation_list):
                    j = j-1
                    title = ws.cell(column = 2, row = j).value

                if title not in titles and title != None:
                    titles.append(title)

                parameters.append(title)





    citation_titles[citation_NO] = parameters

    ws2.cell(row = i+5, column = 1).value = citation_NO

    for value_no, value in enumerate(parameters):
        ws2.cell(row=i + 5, column=2 + value_no).value = value









minimum_items = 3
maximum_items = 6

all_combinations = []
for r_value in range(minimum_items, maximum_items):
    comb = list(combinations(titles, r_value))
    all_combinations.append(comb)


row_no = 6
column_no = 1
for comb1 in all_combinations:
    for index, comb in enumerate(comb1):

        common_citations = []
        for key, value in citation_titles.items():
            result =  all(elem in value  for elem in comb)
            if result:
                common_citations.append(key)




        if len(common_citations) >= 2:
            print(common_citations, comb)

            for index1, value in enumerate(comb):
                ws3.cell(column=index1 + 1, row= row_no).value = value
            row_no += 1


            for index1, value in enumerate(common_citations):
                ws3.cell(column=column_no, row= row_no).value = value
                row_no += 1


            # ws3.cell(column=index1 + 1, row=index + row_no).value = value








wb.save("Paper.xlsx")
