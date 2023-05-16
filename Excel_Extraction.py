import openpyxl
import os
from Basic_Info import *
import datetime
import numpy as np


now = datetime.datetime.now()
dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
def List_Intersection(a, b):
        return [x for x in a if x in b]
def List_Subtractor(a, b):
        return [x for x in a if x not in b]




Input_file = os.path.join(os.getcwd(), "Data", "OpenSEES Input.xlsx")

wb = openpyxl.load_workbook(Input_file, data_only=True)
ws_SB = wb["Section Builder"]
ws_MB = wb["Model Builder"]
ws_GA = wb["Gravity Analysis"]
ws_MA = wb["Modal Analysis"]
ws_TH = wb["TH Analysis"]


##Section Builder
#Row Data






#Modelling Data
GI_row = ws_SB.cell(column = 29, row=3).value
MD_row = ws_SB.cell(column = 29, row=4).value
CM1_row = ws_SB.cell(column = 29, row=5).value
CM2_row = ws_SB.cell(column = 29, row=6).value
SM_row = ws_SB.cell(column = 29, row=7).value
BD_row = ws_SB.cell(column = 29, row=8).value
CD_row = ws_SB.cell(column = 29, row=9).value

def Basic_Parameters(Material_Type, CocnreteUncon_Type, ConcreteCon_Type, Steel_Type):
        # Material Definitions - Concrete and Steel
        column = 3 + Material_Type
        f_c_1 = ws_SB.cell(column=column, row=MD_row + 1).value
        f_c_2 = ws_SB.cell(column=column, row=MD_row + 2).value
        eps_c = ws_SB.cell(column=column, row=MD_row + 3).value
        eps_u = ws_SB.cell(column=column, row=MD_row + 4).value
        f_y = ws_SB.cell(column=column, row=MD_row + 5).value
        E_s = ws_SB.cell(column=column, row=MD_row + 6).value
        Materials_Data = [f_c_1*MPa ,  f_c_2* MPa,  eps_c ,  eps_u ,   f_y*MPa ,    E_s *GPa ]



        # Confined Concrete Material
        column = 3 + ConcreteCon_Type
        ID = ws_SB.cell(column=column, row=CM2_row + 3).value
        matTag = int(ws_SB.cell(column=column, row=CM2_row + 4).value)
        fpc = ws_SB.cell(column=column, row=CM2_row + 5).value
        epsc0 = ws_SB.cell(column=column, row=CM2_row + 6).value
        fpcu = ws_SB.cell(column=column, row=CM2_row + 7).value
        epsU = ws_SB.cell(column=column, row=CM2_row + 8).value
        lamda = ws_SB.cell(column=column, row=CM2_row + 9).value
        ft = ws_SB.cell(column=column, row=CM2_row + 10).value
        Ets = ws_SB.cell(column=column, row=CM2_row + 11).value
        ConcreteCon_Data = [ID, matTag, fpc * MPa, epsc0, fpcu * MPa, epsU, lamda, ft * MPa, Ets * MPa ]


        # UnConfined Concrete Material
        column = 3 + CocnreteUncon_Type
        ID = ws_SB.cell(column=column, row=CM1_row + 3).value
        matTag = int(ws_SB.cell(column=column, row=CM1_row + 4).value)
        fpc = ws_SB.cell(column=column, row=CM1_row + 5).value
        epsc0 = ws_SB.cell(column=column, row=CM1_row + 6).value
        fpcu = ws_SB.cell(column=column, row=CM1_row + 7).value
        epsU = ws_SB.cell(column=column, row=CM1_row + 8).value
        lamda = ws_SB.cell(column=column, row=CM1_row + 9).value
        ft = ws_SB.cell(column=column, row=CM1_row + 10).value
        Ets = ws_SB.cell(column=column, row=CM1_row + 11).value
        ConcreteUnCon_Data= [ID, matTag, fpc * MPa, epsc0, fpcu * MPa, epsU, lamda, ft * MPa, Ets * MPa]



        # Steel Material
        column = 3 + Steel_Type
        ID = ws_SB.cell(column=column, row=SM_row + 2).value
        matTag = int(ws_SB.cell(column=column, row=SM_row + 3).value)
        Fy = ws_SB.cell(column=column, row=SM_row + 4).value
        E0 = ws_SB.cell(column=column, row=SM_row + 5).value
        b = ws_SB.cell(column=column, row=SM_row + 6).value
        R0 = ws_SB.cell(column=column, row=SM_row + 7).value
        cR1 = ws_SB.cell(column=column, row=SM_row + 8).value
        cR2 = ws_SB.cell(column=column, row=SM_row + 9).value
        Steel_Data = [ID, matTag, Fy * MPa, E0 * GPa, b, R0, cR1, cR2 ]

        return Materials_Data, ConcreteUnCon_Data, ConcreteCon_Data,  Steel_Data

def Beam_Column_Data(Section_Type, column):
        if Section_Type == "Beam":

                sec_tag = int(ws_SB.cell(column = column, row =BD_row +  5).value)
                H = ws_SB.cell(column = column, row =BD_row +   6).value
                B = ws_SB.cell(column = column, row =BD_row +   7).value
                cover_H = ws_SB.cell(column = column, row =BD_row +   8).value
                cover_B = ws_SB.cell(column = column, row =BD_row + 9).value
                core_tag =  int(ws_SB.cell(column = column, row =BD_row +10).value)
                cover_tag = int(ws_SB.cell(column = column, row =BD_row +11).value)
                steel_tag = int(ws_SB.cell(column = column, row =BD_row +12).value)
                n_bars_top = ws_SB.cell(column = column, row =BD_row + 13).value
                bar_area_top = ws_SB.cell(column = column, row =BD_row + 14).value
                n_bars_bot = ws_SB.cell(column = column, row =BD_row + 15).value
                bar_area_bot = ws_SB.cell(column = column, row =BD_row + 16).value
                n_bars_int_tot = ws_SB.cell(column = column, row =BD_row + 17).value
                bar_area_int = ws_SB.cell(column = column, row =BD_row +  18).value
                nf_core_y = ws_SB.cell(column = column, row =BD_row +  19).value
                nf_core_z = ws_SB.cell(column = column, row =BD_row +  20).value
                nf_cover_y = ws_SB.cell(column = column, row =BD_row +  21).value
                nf_cover_z = ws_SB.cell(column = column, row =BD_row +  22).value

                Sectional_Data = [sec_tag ,H ,B ,cover_H ,cover_B ,core_tag ,cover_tag ,steel_tag ,n_bars_top ,bar_area_top ,
                                  n_bars_bot ,bar_area_bot ,n_bars_int_tot ,bar_area_int ,
                                  nf_core_y ,nf_core_z ,nf_cover_y ,nf_cover_z]

                return Sectional_Data
        elif  Section_Type == "Column":
                sec_tag = int(ws_SB.cell(column=column, row=CD_row + 5).value)
                H = ws_SB.cell(column=column, row=CD_row + 6).value
                B = ws_SB.cell(column=column, row=CD_row + 7).value
                cover_H = ws_SB.cell(column=column, row=CD_row + 8).value
                cover_B = ws_SB.cell(column=column, row=CD_row + 9).value
                core_tag =  int(ws_SB.cell(column=column, row=CD_row + 10).value)
                cover_tag = int(ws_SB.cell(column=column, row=CD_row + 11).value)
                steel_tag = int(ws_SB.cell(column=column, row=CD_row + 12).value)
                n_bars_top = ws_SB.cell(column=column, row=CD_row + 13).value
                bar_area_top = ws_SB.cell(column=column, row=CD_row + 14).value
                n_bars_bot = ws_SB.cell(column=column, row=CD_row + 15).value
                bar_area_bot = ws_SB.cell(column=column, row=CD_row + 16).value
                n_bars_int_tot = ws_SB.cell(column=column, row=CD_row + 17).value
                bar_area_int = ws_SB.cell(column=column, row=CD_row + 18).value
                nf_core_y = ws_SB.cell(column=column, row=CD_row + 19).value
                nf_core_z = ws_SB.cell(column=column, row=CD_row + 20).value
                nf_cover_y = ws_SB.cell(column=column, row=CD_row + 21).value
                nf_cover_z = ws_SB.cell(column=column, row=CD_row + 22).value
                Sectional_Data = [sec_tag, H, B, cover_H, cover_B, core_tag, cover_tag, steel_tag, n_bars_top, bar_area_top,
                                  n_bars_bot, bar_area_bot, n_bars_int_tot, bar_area_int,
                                  nf_core_y, nf_core_z, nf_cover_y, nf_cover_z]

                return Sectional_Data

def Sectional_Info(Section_Type, Type):
        # _________________Eg: "Beam", "Type1"
        Materials_Data = []
        ConcreteCon_Data = []
        ConcreteUnCon_Data = []
        Steel_Data = []
        Sectional_Data = []


        if Section_Type == "Beam":
                column = 3 + int(Type[-1])
                Material_Type = int(ws_SB.cell(column=column, row=BD_row + 1).value[-1])
                CocnreteUncon_Type= int(ws_SB.cell(column=column, row=BD_row + 2).value[-1])
                ConcreteCon_Type = int(ws_SB.cell(column=column, row=BD_row + 3).value[-1])
                Steel_Type = int(ws_SB.cell(column=column, row=BD_row + 4).value[-1])
                Materials_Data,  ConcreteUnCon_Data, ConcreteCon_Data, Steel_Data = Basic_Parameters(
                Material_Type,  CocnreteUncon_Type ,ConcreteCon_Type, Steel_Type)

                Sectional_Data = Beam_Column_Data("Beam", column)

        elif Section_Type == "Column":
                column = 3 + int(Type[-1])
                Material_Type = int(ws_SB.cell(column=column, row=CD_row + 1).value[-1])
                CocnreteUncon_Type = int(ws_SB.cell(column=column, row=CD_row + 2).value[-1])
                ConcreteCon_Type = int(ws_SB.cell(column=column, row=CD_row + 3).value[-1])
                Steel_Type = int(ws_SB.cell(column=column, row=CD_row + 4).value[-1])

                Materials_Data, ConcreteUnCon_Data, ConcreteCon_Data, Steel_Data = Basic_Parameters(
                        Material_Type, CocnreteUncon_Type, ConcreteCon_Type, Steel_Type)

                Sectional_Data = Beam_Column_Data("Column", column)

        return Materials_Data, ConcreteUnCon_Data, ConcreteCon_Data, Steel_Data, Sectional_Data







#Gravity Analysis Data
RR_row = ws_GA.cell(column = 29, row=3).value
AOR_row = ws_GA.cell(column = 29, row=4).value

def GravityA_Data():
        Recorder_No = int(ws_GA.cell(column = 6, row=RR_row - 1).value)
        Recorder_Data = []
        for column in range(4, 4+Recorder_No):
                Records = []
                for row in range (RR_row + 1, RR_row + 6):
                        Records.append(ws_GA.cell(column = column, row=row).value)

                Recorder_Data.append(Records)


        Analysis_Options = []
        for row in range(AOR_row+1, AOR_row + 10):
                Analysis_Options.append(ws_GA.cell(column = 4, row=row).value)

        return Recorder_Data, Analysis_Options




#Modal Analysis Data
MRR_row = ws_MA.cell(column = 29, row=3).value
MAOR_row = ws_MA.cell(column = 29, row=4).value
def Modal_Data():
        Recorder_No = int(ws_MA.cell(column = 6, row=MRR_row - 1).value)
        Recorder_Data = []
        for column in range(4, 4+Recorder_No):
                Records = []
                for row in range (MRR_row + 1, MRR_row + 6):
                        Records.append(ws_MA.cell(column = column, row=row).value)

                Recorder_Data.append(Records)


        Analysis_Options = []
        for row in range(MAOR_row+1, MAOR_row + 11):
                Analysis_Options.append(ws_MA.cell(column = 4, row=row).value)


        return Recorder_Data, Analysis_Options






#Time History Analysis Data
TRR_row = ws_TH.cell(column = 29, row=3).value
TAOR_row = ws_TH.cell(column = 29, row=4).value
TEQ_row =  ws_TH.cell(column = 29, row=5).value
def TH_Data():
        Recorder_No = int(ws_TH.cell(column = 6, row=TRR_row - 1).value)
        Recorder_Data = []
        for column in range(4, 4+Recorder_No):
                Records = []
                for row in range (TRR_row + 1, TRR_row + 6):
                        Records.append(ws_TH.cell(column = column, row=row).value)

                Recorder_Data.append(Records)


        Analysis_Options = []
        for row in range(TAOR_row+1, TAOR_row + 9):
                Analysis_Options.append(ws_TH.cell(column = 4, row=row).value)


        Earthquake_Info = []
        Earthquake_Nos = ws_TH.cell(column=6, row=TEQ_row-1).value
        for column in range(4, Earthquake_Nos + 4):
                EQs = []
                for row in range(TEQ_row+1, TEQ_row + 16):
                        value = ws_TH.cell(column=column, row=row).value
                        EQs.append(value)


                Scaling_Nos = ws_TH.cell(column=column, row=TEQ_row + 16).value
                Scalings = []
                for row in range(TEQ_row+17, TEQ_row + 17 + Scaling_Nos):
                        value = ws_TH.cell(column=column, row=row).value
                        Scalings.append(value)
                EQs.append(Scalings)

                Earthquake_Info.append(EQs)


        return Recorder_Data, Analysis_Options, Earthquake_Info







#Model Data
MB_GI_row = ws_MB.cell(column = 29, row=3).value
MB_Type1_row = ws_MB.cell(column = 29, row=4).value
MB_Type2_row = ws_MB.cell(column = 29, row=5).value
MB_Type3_row = ws_MB.cell(column = 29, row=6).value
MB_Type4_row = ws_MB.cell(column = 29, row=7).value
Model_Type = ws_MB.cell(column = 6, row=MB_GI_row + 2).value


def Model_Info():
    if Model_Type == "Type 1":
        XBays_No = int(ws_MB.cell(column = 6, row=MB_Type1_row + 1).value)
        YBays_No = int(ws_MB.cell(column = 6, row=MB_Type1_row + 2).value)
        Storey_No= int(ws_MB.cell(column = 6, row=MB_Type1_row + 3).value)

        XSpacing = []
        YSpacing = []
        ZSpacing = []
        XMass = []
        YMass = []
        ZMass = []



        for column in range(3, XBays_No + 3):
            XSpacing.append(float(ws_MB.cell(column = column, row=MB_Type1_row + 6).value))
        for column in range(3, YBays_No + 3):
            YSpacing.append(float(ws_MB.cell(column = column, row=MB_Type1_row + 7).value))
        for column in range(3, Storey_No + 4):
                if column > 3:
                    ZSpacing.append(float(ws_MB.cell(column = column, row=MB_Type1_row + 10).value))
                XMass.append(float(ws_MB.cell(column = column, row=MB_Type1_row + 11).value))
                YMass.append(float(ws_MB.cell(column = column, row=MB_Type1_row + 12).value))
                ZMass.append(float(ws_MB.cell(column = column, row=MB_Type1_row + 13).value))



        #Nodes Data
        Nodes_Name = []
        Node_NameCoOrd = {}

        for z in range(1, len(ZSpacing) + 2):
                for x in range(1, len(XSpacing) + 2):
                        for y in range(1, len(YSpacing) + 2):
                                Node_Name = int(str(z) + str(x) + str(y))

                                XR = [XSpacing[i] for i in range(x-1)]
                                YR = [YSpacing[i] for i in range(y-1)]
                                ZR = [ZSpacing[i] for i in range(z-1)]

                                X_CoOrd = sum(XR)
                                Y_CoOrd = sum(YR)
                                Z_CoOrd = sum(ZR)

                                Co_Ord = [X_CoOrd, Y_CoOrd, Z_CoOrd]

                                Nodes_Name.append(Node_Name)
                                Node_NameCoOrd[Node_Name] = Co_Ord

        # Info Writer
        ModalInfo_Writer(f" ____________________Grid Info ____________________")
        ModalInfo_Writer("XSpacing", Data=XSpacing)
        ModalInfo_Writer("YSpacing", Data=YSpacing)
        ModalInfo_Writer("ZSpacing", Data=ZSpacing)
        ModalInfo_Writer("XMass", Data=XMass)
        ModalInfo_Writer("YMass", Data=YMass)
        ModalInfo_Writer("ZMass", Data=ZMass)


        return Nodes_Name, Node_NameCoOrd, XSpacing, YSpacing, ZSpacing, XMass, YMass, ZMass







#Model Info Writer to txt file handler
Output_txt = os.path.join(os.getcwd(), ws_MB.cell(column = 6, row=MB_GI_row + 1).value)


def ModalInfo_Writer(Heading = None, Data=None, Nth_writer = 1):
        def Write(object, text):
                if text is not None:
                        object.write(str(text))
        if Nth_writer == 0:
                Output_File = open(Output_txt, "a+")
                string_1 = f"\n\n____________________Created On: {dt_string}____________________"
                string_2 = f"_________________________Author: Nolaraj Poudel________________________\n\n"
                Output_File.write(f"{string_1}\n")
                Output_File.write(f"{string_2}\n\n")


        else:

                Output_File = open(Output_txt, "a")
                Output_File.write(f"{Heading}\n")


                Data_Dimension = len(np.array(Data).shape)

                if Data_Dimension == 0:
                        Write(Output_File, Data)
                        Write(Output_File, f"\n")

                elif Data_Dimension == 1:
                        for data in Data:
                                Write(Output_File,  f"{data}\t")
                        Write(Output_File, f"\n")

                else:
                        for datas in Data:
                                for data in datas:
                                        Write(Output_File, f"{data}\t")
                                Write(Output_File, f"\n")


                if Data is not None:
                        Write(Output_File, f"\n")
























